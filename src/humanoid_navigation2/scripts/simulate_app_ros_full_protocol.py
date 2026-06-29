#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""APP-ROS 全功能协议模拟验收。

本脚本把当前已经实现的三个核心能力串起来：
1. 静态核对 APP 文档和 Excel 是否包含当前源码需要的关键 JSON 字段。
2. 动态运行多地图 switch_map 语义模拟，验证 switch_map -> map_ready。
3. 动态运行 route task 语义模拟，验证 start/pause/resume/stop/jump/broadcast 等核心事件。

注意：这仍是协议级和状态机级模拟，不替代真实 hall1 地图上的 RO/OP 定位和实机导航测试。
"""

from __future__ import annotations

import argparse
import os
import signal
import subprocess
import sys
import time
from pathlib import Path
from typing import Iterable, List


WORKSPACE_ROOT = Path(__file__).resolve().parents[3]
DOC_PATH = WORKSPACE_ROOT / "src/humanoid_navigation2/docs/APP侧多地图一期改造方案.md"
EXCEL_PATH = Path("/home/ubuntu/下载/功能指令库 (4).xlsx")


def run_command(command: List[str], title: str, env: dict | None = None, timeout: float = 60.0) -> None:
    """运行子命令，失败时抛出带上下文的异常。"""
    print(f"\n=== {title} ===")
    print(" ".join(command))
    completed = subprocess.run(
        command,
        cwd=str(WORKSPACE_ROOT),
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )
    print(completed.stdout)
    if completed.returncode != 0:
        raise RuntimeError(f"{title} 失败，退出码={completed.returncode}")


def read_excel_text(path: Path) -> str:
    """读取 Excel 全部单元格文本，用于协议字段静态核对。"""
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - 现场缺依赖时给出清晰错误
        raise RuntimeError(f"缺少 openpyxl，无法读取 Excel: {exc}") from exc

    workbook = load_workbook(path, data_only=True)
    chunks: List[str] = []
    for sheet in workbook.worksheets:
        chunks.append(f"[[sheet:{sheet.title}]]")
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    chunks.append(str(cell.value))
    return "\n".join(chunks)


def require_tokens(name: str, text: str, tokens: Iterable[str]) -> None:
    """确认文本包含必要协议字段。"""
    missing = [token for token in tokens if token not in text]
    if missing:
        raise RuntimeError(f"{name} 缺少关键协议字段: {missing}")
    print(f"[OK] {name} 关键字段齐全")


def validate_protocol_artifacts() -> None:
    """静态核对文档和 Excel 是否覆盖全功能协议字段。"""
    required_tokens = [
        "map_management",
        "get_map_list",
        "get_current_map",
        "switch_map",
        "map_response",
        "map_status",
        "waypoint_management",
        "set_waypoint",
        "update_waypoint",
        "clear_waypoints",
        "waypoints_data",
        "waypoints_revision",
        "waypoints_revisions_by_map",
        "navigation_control",
        "start_route_task",
        "pause_route_task",
        "resume_route_task",
        "stop_route_task",
        "jump_to_waypoint",
        "broadcast_finished",
        "route_waypoint_ids",
        "route_waypoints",
        "navigation_command_result",
        "broadcast_requested",
        "waypoint_passed",
        "jump_updated",
        "task_waypoint_completed",
        "route_task_completed",
        "navigation_failed",
        "obstacle_wait",
        "\"map_id\"",
    ]

    doc_text = DOC_PATH.read_text(encoding="utf-8")
    require_tokens("APP 多地图开发文档", doc_text, required_tokens)

    if not EXCEL_PATH.exists():
        raise RuntimeError(f"Excel 协议表不存在: {EXCEL_PATH}")
    excel_text = read_excel_text(EXCEL_PATH)
    require_tokens("功能指令库 Excel", excel_text, required_tokens)

    # 额外核对 APP 网关入口层源码确实支持这三类 APP data_type。
    websocket_source = (
        WORKSPACE_ROOT / "src/humanoid_app_gateway_runtime/src/app_gateway_node.cpp"
    ).read_text(encoding="utf-8")
    require_tokens(
        "app_gateway_node 转发入口",
        websocket_source,
        [
            '"waypoint_management"',
            '"navigation_control"',
            '"map_management"',
        ],
    )

    router_source = (
        WORKSPACE_ROOT / "src/humanoid_app_gateway_runtime/src/business_command_router.cpp"
    ).read_text(encoding="utf-8")
    require_tokens(
        "business_command_router 字段转发",
        router_source,
        [
            '"map_id"',
            '"route_waypoint_ids"',
            '"waypoints_revision"',
            '"target_map_id"',
        ],
    )


def ros_env(domain_id: int) -> dict:
    """构造隔离 ROS_DOMAIN_ID 的环境。"""
    env = os.environ.copy()
    env["ROS_DOMAIN_ID"] = str(domain_id)
    env["FASTDDS_BUILTIN_TRANSPORTS"] = "UDPv4"
    return env


def run_multimap_sim(domain_id: int) -> None:
    """运行多地图 switch_map 语义模拟。"""
    run_command(
        ["python3", "src/humanoid_navigation2/scripts/simulate_multimap_switch_semantics.py"],
        title=f"多地图 switch_map 语义模拟 ROS_DOMAIN_ID={domain_id}",
        env=ros_env(domain_id),
        timeout=30.0,
    )


def run_route_task_sim(domain_id: int) -> None:
    """启动真实 navigation_state_manager 后运行路线任务全场景模拟。"""
    env = ros_env(domain_id)
    process = subprocess.Popen(
        ["ros2", "run", "humanoid_navigation", "navigation_state_manager"],
        cwd=str(WORKSPACE_ROOT),
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        start_new_session=True,
    )
    try:
        # 给节点一点启动时间；真正 ready 由 simulate_route_task_semantics 中的订阅计数等待保证。
        time.sleep(1.0)
        run_command(
            ["python3", "src/humanoid_navigation2/scripts/simulate_route_task_semantics.py", "--scenario", "all"],
            title=f"路线任务 + 跳点 + 播报语义模拟 ROS_DOMAIN_ID={domain_id}",
            env=env,
            timeout=80.0,
        )
    finally:
        if process.poll() is None:
            os.killpg(process.pid, signal.SIGINT)
            try:
                process.wait(timeout=5.0)
            except subprocess.TimeoutExpired:
                os.killpg(process.pid, signal.SIGTERM)
                process.wait(timeout=5.0)
        if process.stdout is not None:
            output = process.stdout.read()
            if output:
                print("--- navigation_state_manager output ---")
                print(output[-6000:])


def main() -> int:
    parser = argparse.ArgumentParser(description="APP-ROS 全功能协议模拟验收")
    parser.add_argument("--ros-domain-id", type=int, default=0, help="指定起始 ROS_DOMAIN_ID")
    args = parser.parse_args()

    base_domain = args.ros_domain_id or (180 + int(time.time()) % 30)
    validate_protocol_artifacts()
    run_multimap_sim(base_domain)
    run_route_task_sim(base_domain + 1)

    print("\nAPP-ROS 全功能协议模拟通过：文档/Excel 关键 JSON 字段、switch_map 切图链路、route task 跳点/播报链路均符合当前源码。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
